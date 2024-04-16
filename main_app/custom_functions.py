from openpyxl import load_workbook
import json, os
from .hod_views import bprint
from django.conf import settings

def excel_to_json(excel_file):
    workbook = load_workbook(excel_file)
    sheet = workbook['Sheet1']
    
    timings = []
    for row in sheet.iter_rows(values_only=True):
        timings.append(row)
    table = {}
    for row in timings[1:]:
        if row[0] is not None:
            day_lower = row[0].lower().capitalize()
            day_schedule = []
            for subject, timing in zip(row[1:], timings[0][1:]):
                if subject and timing:
                    day_schedule.append({
                        "subject": subject.capitalize(),
                        "timing": timing.upper()
                    })
            if day_schedule:
                table[day_lower] = day_schedule
    return json.dumps(table)

def merge_paths(dir, file, file_name=False):
    file = os.path.basename(file)
     
    if file_name:
        return os.path.join(dir, file), file
    return os.path.join(dir, file)